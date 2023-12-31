#!loginusr/bin/python3
# -*- coding: UTF-8 -*-

# Author   : JasonHung
# Date     : 20221102
# Update   : 20230720
# Function : otsuka factory work time record

import pymysql , logging , time , re , requests , json , calendar , csv , json , openpyxl , pyodbc , sys
import matplotlib.pyplot as plt
import mplcursors
from io import BytesIO
import base64
from fpdf import *
from control.config import *
from openpyxl.styles import Font , PatternFill , Alignment


import io
import base64
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas

########################################################################################################################################
#
# web_cloud_dao
#
########################################################################################################################################
class web_cloud_dao:

    ########
    # log
    ########
    log_format = "%(asctime)s %(message)s"
    logging.basicConfig(format=log_format , level=logging.INFO , datefmt="%Y-%m-%d %H:%M:%S")
    #logging.disable(logging.INFO)


    def generate_plot():
        # 创建一个简单的 Matplotlib 图表
        fig, ax = Figure(), FigureCanvas(Figure())
        ax.plot([1, 2, 3, 4, 5], [2, 4, 6, 8, 10])
        return fig

    ##########################
    # show_day_money_detail
    ##########################
    def show_day_money_detail(self , year , month):
        
        self.__connect__()
        
        try:
            month = '0' + month if int(month) < 10 else month

            year_month_sql = f"SELECT day_r_date , day_r_month , day_r_day FROM `day_money` WHERE day_r_year='2023' and day_r_month='09' group by day_r_day order by day_r_day desc"
            name_sql       = f"select a.a_name , b.employee_eng_name , a.day_r_date , a.day_t_money from `day_money` a left join hr_a b on a.a_name = b.employee_name  WHERE  day_r_year='2023' and day_r_month='09' and day_r_month != '9/' order by day_r_day desc"

            self.sql = f"select a_name , day_r_month from `day_money` WHERE  day_r_year='{str(year)}' and day_r_month='{str(month)}' and day_r_month != '9/' order by day_r_day desc" 
            self.curr.execute(self.sql)
            self.res = self.curr.fetchall()
            month    = []

            for val in self.res:
                month.append(val[0]) 

            return month

        except Exception as e:
            logging.error('< Error > show_day_money_detail : ' + str(e))

        finally:
            self.__disconnect__()

    ################################
    # show_day_money_detail_money
    ################################
    def show_day_money_detail_money(self , year , month):
        
        self.__connect__()
        
        try:
            month = '0' + month if int(month) < 10 else month

            name_sql = f"select a_name , e_name from day_money where day_r_year='{year}' and day_r_month='{month}' group by a_name  order by day_r_day asc"
            self.curr.execute(name_sql)
            name_res = self.curr.fetchall()

            for name_val in name_res:
             
                money_sql = f"select day_t_money from day_money where day_r_year='{year}' and day_r_month='{month}' and a_name='{name}' order by day_r_day asc"
                self.curr.execute(money_sql)
                money_res = self.curr.fetchall()

                return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_detail_money : ' + str(e))

        finally:
            self.__disconnect__()

    ################################################
    # show_day_money_parking_fee_detail_day_total
    ################################################
    def show_day_money_parking_fee_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_parking_fee where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_parking_fee_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ############################################
    # show_day_money_over_traffic_detail_day_total
    ############################################
    def show_day_money_over_traffic_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_over_traffic where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_over_traffic_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ############################################
    # show_day_money_traffic_detail_day_total
    ############################################
    def show_day_money_traffic_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_traffic where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_traffic_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ##########################################
    # show_day_money_tolls_detail_day_total
    ##########################################
    def show_day_money_tolls_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_tolls where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_tolls_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ##########################################
    # show_day_money_trick_detail_day_total
    ##########################################
    def show_day_money_trick_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_trick where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_trick_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    #########################################
    # show_day_money_taxi_detail_day_total
    #########################################
    def show_day_money_taxi_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_taxi where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_taxi_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    #########################################
    # show_day_money_stay_detail_day_total
    #########################################
    def show_day_money_stay_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_stay where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_stay_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    #########################################
    # show_day_money_other_detail_day_total
    ##########################################
    def show_day_money_other_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_other where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_other_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    #######################################
    # show_day_money_oil_detail_day_total
    ########################################
    def show_day_money_oil_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_oil where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_oil_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ####################################
    # show_day_money_detail_day_total
    ####################################
    def show_day_money_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()
    
    ###########################################
    # show_day_money_parking_fee_detail_name
    ###########################################
    def show_day_money_parking_fee_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_parking_fee where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/parking_fee_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/parking_fee_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/parking_fee_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_parking_fee where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_parking_fee_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    ######################################
    # show_day_money_over_traffic_detail_name
    ######################################
    def show_day_money_over_traffic_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_over_traffic where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/over_traffic_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/over_traffic_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/over_traffic_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_over_traffic where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_over_traffic_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    ######################################
    # show_day_money_traffic_detail_name
    ######################################
    def show_day_money_traffic_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_traffic where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/traffic_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/traffic_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/traffic_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_traffic where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_traffic_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # show_day_money_tolls_detail_name
    #####################################
    def show_day_money_tolls_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_tolls where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/tolls_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/tolls_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/tolls_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_tolls where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_tolls_detail_name (日當 過路費) : ' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # show_day_money_trick_detail_name
    #####################################
    def show_day_money_trick_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_trick where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/trick_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/trick_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/trick_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_trick where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_trick_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    ####################################
    # show_day_money_taxi_detail_name
    ####################################
    def show_day_money_taxi_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_taxi where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/taxi_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/taxi_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/taxi_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_taxi where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_taxi_detail_name : (日當  計程車)' + str(e))

        finally:
            self.__disconnect__()

    ####################################
    # show_day_money_stay_detail_name
    ####################################
    def show_day_money_stay_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_stay where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/stay_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/stay_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/stay_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_stay where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_stay_detail_name : (日當 住宿)' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # show_day_money_other_detail_name
    #####################################
    def show_day_money_other_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_other where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/other_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/other_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/other_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_other where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_other_detail_name : (日當  其他)' + str(e))

        finally:
            self.__disconnect__()

    ###################################
    # show_day_money_oil_detail_name
    ###################################
    def show_day_money_oil_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_oil where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/oil_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/oil_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/oil_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_oil where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_oil_detail_name : (日當 油票)' + str(e))

        finally:
            self.__disconnect__()

    ##############################
    # show_day_money_detail_name
    ##############################
    def show_day_money_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('< Error > show_day_money_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    ##############################
    # show_day_money_detail_day
    ##############################
    def show_day_money_detail_day(self , year , month):
        
        self.__connect__()
        
        try:
            month = '0' + month if int(month) < 10 else month

            # all day by month 
            day_sql = f"select day_r_month , day_r_day from day_money where day_r_year='{year}' and day_r_month='{month}' group by day_r_day order by day_r_day asc"
            self.curr.execute(day_sql)
            day_res = self.curr.fetchall() 

            return day_res

        except Exception as e:
            logging.error('< Error > show_day_money_detail_day : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # show_computer_chi_name
    ###########################
    def show_computer_chi_name(self , d_name):
        
        self.__connect__()
        
        try:
            # device chinese name
            d_sql = f"select employee_name from hr_a where d_name='{d_name}'"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchone() 

            return d_res

        except Exception as e:
            logging.error('< Error > show_computer_chi_name : ' + str(e))

        finally:
            self.__disconnect__()

    ##############################
    # sensor_position_detail
    ##############################
    def sensor_position_detail(self , sensor):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user by s_number
            d_sql  = f"SELECT r_time , val_1 , val_2 from {r_date} where s_kind='{sensor}' order by no desc limit 0,1"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > sensor_position_detail : ' + str(e))

        finally:
            self.__disconnect4__()

    ##############################
    # computer_s_number_detail
    ##############################
    def computer_s_number_detail(self):
        
        self.__connect__()
        
        try:
            # all device user by s_number
            d_sql  = f"SELECT s_number , d_name , d_status , r_date FROM (SELECT s_number , d_name , d_status , r_date, ROW_NUMBER() OVER (PARTITION BY s_number ORDER BY r_date DESC) AS rn FROM device_list) AS subquery WHERE rn = 1 ORDER BY r_date DESC"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > computer_s_number_detail : ' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # search_show_computer_user_detail
    #####################################
    def search_show_computer_user_detail(self , s_number):
        
        self.__connect__()
        
        try:
            # all device user 
            d_sql = f"select distinct d_name from device_list where s_number='{s_number}' order by l_activity desc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > show_computer_user_detail : ' + str(e))

        finally:
            self.__disconnect__()

    ######################################
    # show_factory_monitor_position_img
    ######################################
    def show_factory_monitor_position_img(self):
        
        self.__connect4__()
        try:
            # record time
            now_month = time.strftime("%Y_%m" , time.localtime()) 
            
            # all device position
            d_sql = f"select distinct b.d_name , b.d_c_name from {now_month} a left join monitor_device b on a.s_kind=b.d_name where a.s_kind != 'I6-1' and a.s_kind !='I6-2' order by b.d_name asc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x = [row[0] for row in d_res]  # x轴数据
            y1 = [row[1] for row in d_res]  # 第一条线的y轴数据
            y2 = [row[2] for row in d_res]  # 第二条线的y轴数据

            # 生成多条线图
            plt.plot(x, y1, label='Line 1')
            plt.plot(x, y2, label='Line 2')
            plt.xlabel('X Label')
            plt.ylabel('Y Label')
            plt.title('MySQL Data Multiple Line Chart')
            plt.legend()

            # 保存图像到内存中
            img = io.BytesIO()
            plt.savefig(img, format='png')
            img.seek(0)

            # 将图像转换为base64编码
            plot_url = base64.b64encode(img.getvalue()).decode()


            return plot_url

        except Exception as e:
            logging.error('< Error > show_factory_monitor_position_img : ' + str(e))

        finally:
            self.__disconnect4__()

    ##################################
    # show_factory_monitor_position
    ##################################
    def show_factory_monitor_position(self):
        
        self.__connect4__()
        try:
            # record time
            now_month = time.strftime("%Y_%m" , time.localtime()) 
            
            # all device position
            d_sql = f"select distinct b.d_name , b.d_c_name from {now_month} a left join monitor_device b on a.s_kind=b.d_name where a.s_kind != 'I6-1' and a.s_kind !='I6-2' order by b.d_name asc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > show_factory_monitor_position : ' + str(e))

        finally:
            self.__disconnect4__()

    ######################################
    # show_factory_monitor_detail_chart
    ######################################
    def show_factory_monitor_detail_chart(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,20"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################
    # show_factory_monitor_detail_rh_img
    #########################################
    def show_factory_monitor_detail_rh_img(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x  = [row[0] for row in d_res]  # x轴数据
            y1 = [float(row[4]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='RH' ,  marker='o', markersize=4)
            axis.set_title('RH')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (%)')
            
            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_rh_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################
    # show_factory_monitor_detail_rh_val
    #########################################
    def show_factory_monitor_detail_rh_val(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select min(val_2) , max(val_2) , ROUND(AVG(val_2),2) from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_rh_val : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################
    # show_factory_monitor_detail_temp_val
    #########################################
    def show_factory_monitor_detail_temp_val(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select min(val_1) , max(val_1) , ROUND(AVG(val_1),2) from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_temp_val : ' + str(e))

        finally:
            self.__disconnect4__()

    #############################################
    # show_factory_monitor_detail_rh_pie_img
    #############################################
    def show_factory_monitor_detail_rh_pie_img(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select val_2 , count(*) from {r_date} where s_kind='{s_kind}' group by val_2 order by val_2 desc limit 0,8"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            val = [row[0] for row in d_res]  # x轴数据
            count = [float(row[1]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            # 创建圆饼图
            
            axis.pie(count, labels=val, autopct='%1.1f%%', startangle=90)
            axis.axis('equal')  # 保证饼图是圆形的
            axis.set_title('RH')

            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_rh_pie_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #############################################
    # show_factory_monitor_detail_temp_pie_img
    #############################################
    def show_factory_monitor_detail_temp_pie_img(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select val_1 , count(*) from {r_date} where s_kind='{s_kind}' group by val_1 order by val_1 desc limit 0,8"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            val = [row[0] for row in d_res]  # x轴数据
            count = [float(row[1]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            # 创建圆饼图
            
            axis.pie(count, labels=val, autopct='%1.1f%%', startangle=90)
            axis.axis('equal')  # 保证饼图是圆形的
            axis.set_title('Temp')

            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_temp_pie_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################
    # show_factory_monitor_detail_temp_img
    #########################################
    def show_factory_monitor_detail_temp_img(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x  = [row[0] for row in d_res]  # x轴数据
            y1 = [float(row[3]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='Temp' ,  marker='o', markersize=4)
            axis.set_title('Temp')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (°C)')

            # 启用鼠标悬停提示显示数值
            mplcursors.cursor(fig).connect("add", lambda axis: axis.annotation.set_text(f"{axis.target[0]:.2f},{axis.target[1]:.2f}"))
            # 添加提示文本（手动方式）
            #for i, txt in enumerate(y1):
            #    axis.annotate(f'{txt:.2f}', (x[i], y1[i]), textcoords="offset points", xytext=(0,10), ha='center')

            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_temp_img : ' + str(e))

        finally:
            self.__disconnect4__()

    ##############################################
    # show_factory_monitor_detail_temp_rh_img_2
    ##############################################
    def show_factory_monitor_detail_temp_rh_img_2(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x  = [row[0] for row in d_res]  # x轴数据
            y1 = [float(row[4]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='RH' ,  marker='o', markersize=4)
            axis.set_title(s_kind + ' RH')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (%)')

            # 启用鼠标悬停提示显示数值
            mplcursors.cursor(fig).connect("add", lambda axis: axis.annotation.set_text(f"{axis.target[0]:.2f},{axis.target[1]:.2f}"))
            # 添加提示文本（手动方式）
            #for i, txt in enumerate(y1):
            #    axis.annotate(f'{txt:.2f}', (x[i], y1[i]), textcoords="offset points", xytext=(0,10), ha='center')
            
            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_temp_rh_img_2 : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################################
    # show_factory_monitor_detail_warehouse_rh_img
    #########################################################
    def show_factory_monitor_detail_warehouse_rh_img(self):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## S-1
            s_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-1' order by r_time desc limit 0,40"
            self.curr.execute(s_1_sql)
            s_1_res = self.curr.fetchall() 
            ### S-2
            s_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-2' order by r_time desc limit 0,40"
            self.curr.execute(s_2_sql)
            s_2_res = self.curr.fetchall() 
            ### S-3
            s_3_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-3' order by r_time desc limit 0,40"
            self.curr.execute(s_3_sql)
            s_3_res = self.curr.fetchall() 
            ### S-4
            s_4_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-4' order by r_time desc limit 0,40"
            self.curr.execute(s_4_sql)
            s_4_res = self.curr.fetchall() 
            ### S-5
            s_5_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-5' order by r_time desc limit 0,40"
            self.curr.execute(s_5_sql)
            s_5_res = self.curr.fetchall() 
            ### S-6
            s_6_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-6' order by r_time desc limit 0,40"
            self.curr.execute(s_6_sql)
            s_6_res = self.curr.fetchall() 
            ### S-7
            s_7_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-7' order by r_time desc limit 0,40"
            self.curr.execute(s_7_sql)
            s_7_res = self.curr.fetchall() 
            ### S-8
            s_8_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-8' order by r_time desc limit 0,40"
            self.curr.execute(s_8_sql)
            s_8_res = self.curr.fetchall() 
            ### S-9
            s_9_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-9' order by r_time desc limit 0,40"
            self.curr.execute(s_9_sql)
            s_9_res = self.curr.fetchall() 
            ### S-10
            s_10_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-10' order by r_time desc limit 0,40"
            self.curr.execute(s_10_sql)
            s_10_res = self.curr.fetchall() 
            ### S-14
            s_14_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-14' order by r_time desc limit 0,40"
            self.curr.execute(s_14_sql)
            s_14_res = self.curr.fetchall() 
            ### S-17
            s_17_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-17' order by r_time desc limit 0,40"
            self.curr.execute(s_17_sql)
            s_17_res = self.curr.fetchall() 
            ### S-18
            s_18_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-18' order by r_time desc limit 0,40"
            self.curr.execute(s_18_sql)
            s_18_res = self.curr.fetchall() 

            x  = [row[0] for row in s_1_res]  # x轴数据
            y1 = [float(row[4]) for row in s_1_res]  # S-1
            y2 = [float(row[4]) for row in s_2_res]  # S-2
            y3 = [float(row[4]) for row in s_3_res]  # S-3
            y4 = [float(row[4]) for row in s_4_res]  # S-4
            y5 = [float(row[4]) for row in s_5_res]  # S-5
            y6 = [float(row[4]) for row in s_6_res]  # S-6
            y7 = [float(row[4]) for row in s_7_res]  # S-7
            y8 = [float(row[4]) for row in s_8_res]  # S-8
            y9 = [float(row[4]) for row in s_9_res]  # S-9
            y10 = [float(row[4]) for row in s_10_res]  # S-10
            y14 = [float(row[4]) for row in s_14_res]  # S-14
            y17 = [float(row[4]) for row in s_17_res]  # S-17
            y18 = [float(row[4]) for row in s_18_res]  # S-14
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='S-1')
            axis.plot(x, y2 , label='S-2')
            axis.plot(x, y3 , label='S-3')
            axis.plot(x, y4 , label='S-4')
            axis.plot(x, y5 , label='S-5')
            axis.plot(x, y6 , label='S-6')
            axis.plot(x, y7 , label='S-7')
            axis.plot(x, y8 , label='S-8')
            axis.plot(x, y9 , label='S-9')
            axis.plot(x, y10 , label='S-10')
            axis.plot(x, y14 , label='S-14')
            axis.plot(x, y17 , label='S-17')
            axis.plot(x, y18 , label='S-18')

            axis.set_title('Warehouse Sensor RH')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (%)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_warehouse_rh_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################################
    # show_factory_monitor_detail_warehouse_temp_img
    #########################################################
    def show_factory_monitor_detail_warehouse_temp_img(self):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## S-1
            s_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-1' order by r_time desc limit 0,40"
            self.curr.execute(s_1_sql)
            s_1_res = self.curr.fetchall() 
            ### S-2
            s_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-2' order by r_time desc limit 0,40"
            self.curr.execute(s_2_sql)
            s_2_res = self.curr.fetchall() 
            ### S-3
            s_3_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-3' order by r_time desc limit 0,40"
            self.curr.execute(s_3_sql)
            s_3_res = self.curr.fetchall() 
            ### S-4
            s_4_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-4' order by r_time desc limit 0,40"
            self.curr.execute(s_4_sql)
            s_4_res = self.curr.fetchall() 
            ### S-5
            s_5_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-5' order by r_time desc limit 0,40"
            self.curr.execute(s_5_sql)
            s_5_res = self.curr.fetchall() 
            ### S-6
            s_6_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-6' order by r_time desc limit 0,40"
            self.curr.execute(s_6_sql)
            s_6_res = self.curr.fetchall() 
            ### S-7
            s_7_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-7' order by r_time desc limit 0,40"
            self.curr.execute(s_7_sql)
            s_7_res = self.curr.fetchall() 
            ### S-8
            s_8_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-8' order by r_time desc limit 0,40"
            self.curr.execute(s_8_sql)
            s_8_res = self.curr.fetchall() 
            ### S-9
            s_9_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-9' order by r_time desc limit 0,40"
            self.curr.execute(s_9_sql)
            s_9_res = self.curr.fetchall() 
            ### S-10
            s_10_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-10' order by r_time desc limit 0,40"
            self.curr.execute(s_10_sql)
            s_10_res = self.curr.fetchall() 
            ### S-14
            s_14_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-14' order by r_time desc limit 0,40"
            self.curr.execute(s_14_sql)
            s_14_res = self.curr.fetchall() 
            ### S-17
            s_17_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-17' order by r_time desc limit 0,40"
            self.curr.execute(s_17_sql)
            s_17_res = self.curr.fetchall() 
            ### S-18
            s_18_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-18' order by r_time desc limit 0,40"
            self.curr.execute(s_18_sql)
            s_18_res = self.curr.fetchall() 

            x  = [row[0] for row in s_1_res]  # x轴数据
            y1 = [float(row[3]) for row in s_1_res]  # S-1
            y2 = [float(row[3]) for row in s_2_res]  # S-2
            y3 = [float(row[3]) for row in s_3_res]  # S-3
            y4 = [float(row[3]) for row in s_4_res]  # S-4
            y5 = [float(row[3]) for row in s_5_res]  # S-5
            y6 = [float(row[3]) for row in s_6_res]  # S-6
            y7 = [float(row[3]) for row in s_7_res]  # S-7
            y8 = [float(row[3]) for row in s_8_res]  # S-8
            y9 = [float(row[3]) for row in s_9_res]  # S-9
            y10 = [float(row[3]) for row in s_10_res]  # S-10
            y14 = [float(row[3]) for row in s_14_res]  # S-14
            y17 = [float(row[3]) for row in s_17_res]  # S-17
            y18 = [float(row[3]) for row in s_18_res]  # S-14
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='S-1')
            axis.plot(x, y2 , label='S-2')
            axis.plot(x, y3 , label='S-3')
            axis.plot(x, y4 , label='S-4')
            axis.plot(x, y5 , label='S-5')
            axis.plot(x, y6 , label='S-6')
            axis.plot(x, y7 , label='S-7')
            axis.plot(x, y8 , label='S-8')
            axis.plot(x, y9 , label='S-9')
            axis.plot(x, y10 , label='S-10')
            axis.plot(x, y14 , label='S-14')
            axis.plot(x, y17 , label='S-17')
            axis.plot(x, y18 , label='S-18')

            axis.set_title('Warehouse Sensor Temp')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (°C)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_warehouse_temp_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################################
    # show_factory_monitor_detail_quality_control_rh_img
    #########################################################
    def show_factory_monitor_detail_quality_control_rh_img(self):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## S-11-1
            s_11_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-11-1' order by r_time desc limit 0,40"
            self.curr.execute(s_11_1_sql)
            s_11_1_res = self.curr.fetchall() 
            ## S-11-2
            s_11_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-11-2' order by r_time desc limit 0,40"
            self.curr.execute(s_11_2_sql)
            s_11_2_res = self.curr.fetchall() 
            ### S-12
            s_12_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-12' order by r_time desc limit 0,40"
            self.curr.execute(s_12_sql)
            s_12_res = self.curr.fetchall() 
            ### S-13
            s_13_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-13' order by r_time desc limit 0,40"
            self.curr.execute(s_13_sql)
            s_13_res = self.curr.fetchall() 
            ### S-15-1
            s_15_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-1' order by r_time desc limit 0,40"
            self.curr.execute(s_15_1_sql)
            s_15_1_res = self.curr.fetchall()
            ### S-15-2
            s_15_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-2' order by r_time desc limit 0,40"
            self.curr.execute(s_15_2_sql)
            s_15_2_res = self.curr.fetchall() 
            ### S-15-3
            s_15_3_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-3' order by r_time desc limit 0,40"
            self.curr.execute(s_15_3_sql)
            s_15_3_res = self.curr.fetchall() 
            ### S-15-4
            s_15_4_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-4' order by r_time desc limit 0,40"
            self.curr.execute(s_15_4_sql)
            s_15_4_res = self.curr.fetchall() 
            ### S-15-5
            s_15_5_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-5' order by r_time desc limit 0,40"
            self.curr.execute(s_15_5_sql)
            s_15_5_res = self.curr.fetchall() 
            ### S-15-6
            s_15_6_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-6' order by r_time desc limit 0,40"
            self.curr.execute(s_15_6_sql)
            s_15_6_res = self.curr.fetchall() 
            ### S-16
            s_16_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-16' order by r_time desc limit 0,40"
            self.curr.execute(s_16_sql)
            s_16_res = self.curr.fetchall() 
            ### S-19
            s_19_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-19' order by r_time desc limit 0,40"
            self.curr.execute(s_19_sql)
            s_19_res = self.curr.fetchall() 

            x  = [row[0] for row in s_11_1_res]  # x轴数据
            y11_1 = [float(row[4]) for row in s_11_1_res]  # S-11-1
            y11_2 = [float(row[4]) for row in s_11_2_res]  # S-11-2
            y12   = [float(row[4]) for row in s_12_res]  # S-12
            y13   = [float(row[4]) for row in s_13_res]  # S-13
            y15_1 = [float(row[4]) for row in s_15_1_res]  # S-15-1
            y15_2 = [float(row[4]) for row in s_15_2_res]  # S-15-2
            y15_3 = [float(row[4]) for row in s_15_3_res]  # S-15-3
            y15_4 = [float(row[4]) for row in s_15_4_res]  # S-15-4
            y15_5 = [float(row[4]) for row in s_15_5_res]  # S-15-5
            y15_6 = [float(row[4]) for row in s_15_6_res]  # S-15-6
            y16   = [float(row[4]) for row in s_16_res]  # S-16
            y19   = [float(row[4]) for row in s_19_res]  # S-19
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y11_1 , label='S-11-1')
            axis.plot(x, y11_2 , label='S-11-2')
            axis.plot(x, y12 , label='S-12')
            axis.plot(x, y13 , label='S-13')
            axis.plot(x, y15_1 , label='S-15-1')
            axis.plot(x, y15_2 , label='S-15-2')
            axis.plot(x, y15_3 , label='S-15-3')
            axis.plot(x, y15_4 , label='S-15-4')
            axis.plot(x, y15_5 , label='S-15-5')
            axis.plot(x, y15_6 , label='S-15-6')
            axis.plot(x, y16 , label='S-16')
            axis.plot(x, y19 , label='S-19')

            axis.set_title('Quality Control Sensor RH')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (%)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_quality_control_rh_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################################
    # show_factory_monitor_detail_quality_control_temp_img
    #########################################################
    def show_factory_monitor_detail_quality_control_temp_img(self):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## S-11-1
            s_11_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-11-1' order by r_time desc limit 0,40"
            self.curr.execute(s_11_1_sql)
            s_11_1_res = self.curr.fetchall() 
            ## S-11-2
            s_11_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-11-2' order by r_time desc limit 0,40"
            self.curr.execute(s_11_2_sql)
            s_11_2_res = self.curr.fetchall() 
            ### S-12
            s_12_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-12' order by r_time desc limit 0,40"
            self.curr.execute(s_12_sql)
            s_12_res = self.curr.fetchall() 
            ### S-13
            s_13_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-13' order by r_time desc limit 0,40"
            self.curr.execute(s_13_sql)
            s_13_res = self.curr.fetchall() 
            ### S-15-1
            s_15_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-1' order by r_time desc limit 0,40"
            self.curr.execute(s_15_1_sql)
            s_15_1_res = self.curr.fetchall()
            ### S-15-2
            s_15_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-2' order by r_time desc limit 0,40"
            self.curr.execute(s_15_2_sql)
            s_15_2_res = self.curr.fetchall() 
            ### S-15-3
            s_15_3_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-3' order by r_time desc limit 0,40"
            self.curr.execute(s_15_3_sql)
            s_15_3_res = self.curr.fetchall() 
            ### S-15-4
            s_15_4_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-4' order by r_time desc limit 0,40"
            self.curr.execute(s_15_4_sql)
            s_15_4_res = self.curr.fetchall() 
            ### S-15-5
            s_15_5_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-5' order by r_time desc limit 0,40"
            self.curr.execute(s_15_5_sql)
            s_15_5_res = self.curr.fetchall() 
            ### S-15-6
            s_15_6_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-6' order by r_time desc limit 0,40"
            self.curr.execute(s_15_6_sql)
            s_15_6_res = self.curr.fetchall() 
            ### S-16
            s_16_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-16' order by r_time desc limit 0,40"
            self.curr.execute(s_16_sql)
            s_16_res = self.curr.fetchall() 
            ### S-19
            s_19_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-19' order by r_time desc limit 0,40"
            self.curr.execute(s_19_sql)
            s_19_res = self.curr.fetchall() 

            x  = [row[0] for row in s_11_1_res]  # x轴数据
            y11_1 = [float(row[3]) for row in s_11_1_res]  # S-11-1
            y11_2 = [float(row[3]) for row in s_11_2_res]  # S-11-2
            y12 = [float(row[3]) for row in s_12_res]  # S-12
            y13 = [float(row[3]) for row in s_13_res]  # S-13
            y15_1 = [float(row[3]) for row in s_15_1_res]  # S-15-1
            y15_2 = [float(row[3]) for row in s_15_2_res]  # S-15-2
            y15_3 = [float(row[3]) for row in s_15_3_res]  # S-15-3
            y15_4 = [float(row[3]) for row in s_15_4_res]  # S-15-4
            y15_5 = [float(row[3]) for row in s_15_5_res]  # S-15-5
            y15_6 = [float(row[3]) for row in s_15_6_res]  # S-15-6
            y16 = [float(row[3]) for row in s_16_res]  # S-16
            y19 = [float(row[3]) for row in s_19_res]  # S-19
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y11_1 , label='S-11-1')
            axis.plot(x, y11_2 , label='S-11-2')
            axis.plot(x, y12 , label='S-12')
            axis.plot(x, y13 , label='S-13')
            axis.plot(x, y15_1 , label='S-15-1')
            axis.plot(x, y15_2 , label='S-15-2')
            axis.plot(x, y15_3 , label='S-15-3')
            axis.plot(x, y15_4 , label='S-15-4')
            axis.plot(x, y15_5 , label='S-15-5')
            axis.plot(x, y15_6 , label='S-15-6')
            axis.plot(x, y16 , label='S-16')
            axis.plot(x, y19 , label='S-19')

            axis.set_title('Quality Control Sensor Temp')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (°C)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_quality_control_temp_img : ' + str(e))

        finally:
            self.__disconnect4__()

    ##############################################
    # show_factory_monitor_detail_temp_rh_img_1
    ##############################################
    def show_factory_monitor_detail_temp_rh_img_1(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x  = [row[0] for row in d_res]  # x轴数据
            y1 = [float(row[3]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='Temp' ,  marker='o', markersize=4)
            axis.set_title(s_kind + ' Temp')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (°C)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail_temp_rh_img_1 : ' + str(e))

        finally:
            self.__disconnect4__()

    ################################
    # show_factory_monitor_detail
    ################################
    def show_factory_monitor_detail(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > show_factory_monitor_detail : ' + str(e))

        finally:
            self.__disconnect4__()

    ##############################
    # show_computer_user_detail
    ##############################
    def show_computer_user_detail(self , d_name):
        
        self.__connect__()
        
        try:
            # all device user 
            d_sql = f"select r_date , d_status , o_name , l_activity , e_ip , i_ip , s_number , s_model , s_manu , registered , cpu_usage , ram_usage , disk_usage from device_list where d_name='{d_name}' order by l_activity desc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > show_computer_user_detail : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##########################
    # show_device_name_list
    ##########################
    def show_device_name_list(self):
        
        self.__connect__()
        
        try:
            # all device user 
            d_sql = f"select r_date , d_name from device_list group by d_name order by r_date desc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('< Error > show_device_name_list : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # bpm_day_money_by_month
    ###########################
    def bpm_day_money_by_month(self , year):
        
        self.__connect__()
        
        try:
            month_sql = f"select day_r_month from `day_money` WHERE  day_r_year='{str(year)}' group by day_r_month order by day_r_month desc" 
            self.curr.execute(month_sql)
            month_res = self.curr.fetchall()

            return month_res

        except Exception as e:
            logging.error('< Error > bpm_day_money_by_month : ' + str(e))

        finally:
            self.__disconnect__()

    ##########################
    # bpm_day_money_by_year
    ##########################
    def bpm_day_money_by_year(self):
        
        self.__connect__()
        
        try:
            self.sql = f"select day_r_year from `day_money` WHERE day_r_year != '112/' group by day_r_year order by day_r_year desc" 
            self.curr.execute(self.sql)
            self.res = self.curr.fetchall()

            return self.res

        except Exception as e:
            logging.error('< Error > bpm_day_money_by_year : ' + str(e))

        finally:
            self.__disconnect__()

    ##################
    # bpm_day_money
    ##################
    def bpm_day_money(self):
        
        self.__connect__()
        
        try:

            self.sql = f"SELECT day_money.r_date , day_money.c_name , hr_a.employee_eng_name , day_money.t_money FROM `day_money` left join hr_a on day_money.c_name = hr_a.employee_name WHERE day_money.r_year='2023' and day_money.r_month='09' order by day_money.r_day desc" 
            self.curr_mssql.execute(self.sql)
            self.res           = self.curr_mssql.fetchall()

            for val in self.res:
                logging.info(f"{val[0]} , {val[1]} , {val[2]} , {val[3]}")


        except Exception as e:
            logging.info('< Error > bpm_day_money : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##################
    # bpm_account_list
    ##################
    def bpm_account_list(self):
        try:
            self.__connect_mssql__()
        except Exception as e:
            logging.info('< Error > bpm_account_list : ' + str(e))
        finally:
            pass

    ########################
    # erp_hr_account_list
    ########################
    def erp_hr_account_list(self):
        try:
            ######################
            #
            # select from MsSQL
            #
            ######################
            if sys.platform.startswith('win'):
                conn_str = f"DRIVER={{SQL Server}};SERVER={otsuka_factory3['host']};DATABASE={otsuka_factory3['db']};UID={otsuka_factory3['user']};PWD={otsuka_factory3['pwd']}"  
            elif sys.platform.startswith('darwin'):
                conn_str = f"DRIVER={{/opt/homebrew/Cellar/msodbcsql17/17.10.5.1/lib/libmsodbcsql.17.dylib}};SERVER={otsuka_factory3['host']};DATABASE={otsuka_factory3['db']};UID={otsuka_factory3['user']};PWD={otsuka_factory3['pwd']}"  
            
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql        = f"SELECT a.EMPID AS EmployeeID, CASE (isnull(a.HECNAME, '')) WHEN '' THEN '' ELSE a.HECNAME END AS EmployeeName, CASE (isnull(a.HEENAME, '')) WHEN '' THEN '' ELSE a.HEENAME END AS EmployeeEnglishName, CASE (isnull(a.LOGIN_ID, '')) WHEN '' THEN '' ELSE a.LOGIN_ID END AS LoginID, a.CPNYID AS CompanyID, a.DEPT_NO AS DepartmentID, '' AS IdentityID, a.SEX, CASE (isnull(a.EMAIL, '')) WHEN '' THEN '' ELSE a.EMAIL END AS Email, CASE (isnull(a.MOBILNO, '')) WHEN '' THEN '' ELSE a.MOBILNO END AS Mobile, SUBSTRING(a.BIRTHDAY, 1, 4) + '/' + SUBSTRING(a.BIRTHDAY, 5, 2) + '/' + SUBSTRING(a.BIRTHDAY, 5, 2) AS Birthday, a.POSSIE AS JobTitleCode, CASE (isnull(b.POS_NAME, '')) WHEN '' THEN '' ELSE b.POS_NAME END AS JobTitleName, CASE (isnull(a.GRADE, '')) WHEN '' THEN '' ELSE a.GRADE END AS JobGrade, CASE (isnull(a.RANK, '')) WHEN '' THEN '' ELSE a.RANK END AS JobRank, '' AS JobCode, '' AS JobType, SUBSTRING(a.INADATE, 1, 4) + '/' + SUBSTRING(a.INADATE, 5, 2) + '/' + SUBSTRING(a.INADATE, 5, 2) AS EnterDate, CASE (isnull(a.PLACE, '')) WHEN '' THEN '' ELSE a.PLACE END AS WorkPlace, '' AS AreaCode, CASE (isnull(a.MOBILNO, '')) WHEN '' THEN '' ELSE a.MOBILNO END AS HomePhone, CASE (isnull(a.EXT, '')) WHEN '' THEN '' ELSE a.EXT END AS OfficePhone, CASE (isnull(a.COMADDR, '')) WHEN '' THEN '' ELSE a.COMADDR END AS Address, '' AS Synopsis FROM dbo.HRUSER AS a LEFT OUTER JOIN dbo.POSITION AS b ON a.POSSIE = b.POSSIE where a.STATE='A'" 
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchall()

            self.__connect__()
            for val in self.res:
                
                s_dep_code_sql = f"select DEP_CODE , DEP_SHORT_NAME from HRUSER_DEPT_BAS where DEP_NO='{val[5]}'"
                self.curr_mssql.execute(s_dep_code_sql)
                res_dep_code = self.curr_mssql.fetchall()

                for dep_val in res_dep_code:

                    ###########################
                    #
                    # check MsSQL hr account 
                    #
                    ###########################
                    s_sql = f"select employee_name from hr_a where employee_name='{val[1]}'"
                    self.curr.execute(s_sql)
                    s_r = self.curr.fetchone()

                    if s_r is None:
                        ######################
                        #
                        # insert into MySQL
                        #
                        ######################
                        sql  = f"insert into hr_a(employee_id , employee_name , employee_eng_name , login_id , company_id , department_id , identity_id , sex , email , mobile , birthday , job_title_code , job_title_name , job_grade , job_rank , job_code , job_type , end_date , work_place , area_code , home_phone , office_phone , addresses , department_code , department_name) "
                        sql += f"value('{val[0]}','{val[1]}','{val[2]}','{val[3]}','{val[4]}','{val[5]}','{val[6]}','{val[7]}','{val[8]}','{val[9]}','{val[10]}','{val[11]}','{val[12]}','{val[13]}','{val[14]}','{val[15]}','{val[16]}','{val[17]}','{val[18]}','{val[19]}','{val[20]}','{val[21]}','{val[22]}','{dep_val[0]}','{dep_val[1]}')"
                        self.curr.execute(sql)
                        self.conn.commit()
                    else:
                        pass
                        #logging.info(f"{s_r[0]} 已存在.")
            
            print('\n')
            logging.info(f"< Msg > HR account , 更新完成.")
            
            self.__disconnect__()

            return self.res
        
        except Exception as e:
            logging.info('< Error > erp_hr_account_list : ' + str(e))

        finally:
            self.curr_mssql.close()
            self.conn_mssql.close()

    ############################
    # department_account_list
    ############################
    def department_account_list(self):
        
        self.__connect__()

        try:
            
            sql  = f"SELECT department_name , department_code , department_id , count(*) FROM `hr_a` group by department_code order by department_name desc" 

            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res
            
        except Exception as e:
            logging.info('< Error > department_account_list : ' + str(e))

        finally:
            self.__disconnect__()
    
    ###########################
    # department_list_detail
    ###########################
    def department_list_detail(self , d_code):
        
        self.__connect__()
        
        try:
            ### connect mysql
            connect_sql = f"select department_name , employee_id , employee_name , department_code from hr_a where department_code='{d_code}' and login_id!='disabled' order by department_name asc"
            self.curr.execute(connect_sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
            
        except Exception as e:
            logging.info('< Error > department_list_detail : ' + str(e))

        finally:
            self.__disconnect__()

    #############################
    # department_no_search_val
    #############################
    def department_no_search_val(self , employee_name):
        
        self.__connect__()
        
        try:
            ### connect mysql
            connect_sql = f"select department_name , department_code from hr_a where employee_name='{employee_name}'"
            self.curr.execute(connect_sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res

            ### connect mssql
            '''
            conn_str        = f"DRIVER={{SQL Server}};SERVER={otsuka_factory3['host']};DATABASE={otsuka_factory3['db']};UID={otsuka_factory3['user']};PWD={otsuka_factory3['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql        = f"SELECT B.DEP_CODE FROM HR_Employee A , HRUSER_DEPT_BAS B WHERE A.DepartmentID = B.DEP_NO AND A.EmployeeName='{employee_name}'"
            self.curr_mssql.execute(self.sql)
            
            self.res        = self.curr_mssql.fetchone()
            self.curr_mssql.commit()
            
            return self.res[0]

            self.curr_mssql.close()
            self.conn_mssql.close()
            '''
            
        except Exception as e:
            logging.info('< Error > department_no_search_vals : ' + str(e))

        finally:
            self.__disconnect__()

    #############################
    # load_work_time_data_list
    #############################
    def load_work_time_data_list(self , e_id , e_name , b_date):
        
        self.__connect__()

        try:
            sql  = f"select normal_time , over_time , availability_time , total_time , b_date , " 
            sql += f"w_s_1 , w_s_1_product , w_s_1_num , w_s_1_normal_time , w_s_1_over_time , w_s_1_avail_time , w_s_1_remark ," 
            sql += f"w_s_2 , w_s_2_product , w_s_2_num , w_s_2_normal_time , w_s_2_over_time , w_s_2_avail_time , w_s_2_remark ," 
            sql += f"w_s_3 , w_s_3_product , w_s_3_num , w_s_3_normal_time , w_s_3_over_time , w_s_3_avail_time , w_s_3_remark ," 
            sql += f"w_s_4 , w_s_4_product , w_s_4_num , w_s_4_normal_time , w_s_4_over_time , w_s_4_avail_time , w_s_4_remark ," 
            sql += f"w_s_5 , w_s_5_product , w_s_5_num , w_s_5_normal_time , w_s_5_over_time , w_s_5_avail_time , w_s_5_remark ," 
            sql += f"w_s_6 , w_s_6_product , w_s_6_num , w_s_6_normal_time , w_s_6_over_time , w_s_6_avail_time , w_s_6_remark ," 
            sql += f"w_s_7 , w_s_7_product , w_s_7_num , w_s_7_normal_time , w_s_7_over_time , w_s_7_avail_time , w_s_7_remark ," 
            sql += f"w_s_8 , w_s_8_product , w_s_8_num , w_s_8_normal_time , w_s_8_over_time , w_s_8_avail_time , w_s_8_remark ," 
            sql += f"w_s_9 , w_s_9_product , w_s_9_num , w_s_9_normal_time , w_s_9_over_time , w_s_9_avail_time , w_s_9_remark ," 
            sql += f"w_s_10 , w_s_10_product , w_s_10_num , w_s_10_normal_time , w_s_10_over_time , w_s_10_avail_time , w_s_10_remark ," 
            sql += f"w_s_11 , w_s_11_product , w_s_11_num , w_s_11_normal_time , w_s_11_over_time , w_s_11_avail_time , w_s_11_remark ," 
            sql += f"w_s_12 , w_s_12_product , w_s_12_num , w_s_12_normal_time , w_s_12_over_time , w_s_12_avail_time , w_s_12_remark " 
            sql += f"from work_time where e_id='{e_id}' and e_name='{e_name}' and b_date='{b_date}'"

            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res

        except Exception as e:
            logging.info('< Error > load_work_time_data_list : ' + str(e))

        finally:
            self.__disconnect__()
    
    ################################
    # load_check_member_data_list3
    ################################
    def load_check_member_data_list3(self , e_name):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 
                
            s_sql2 = f"select p_name , r_time , e_name from card_reader_{now_day} where e_name='{e_name}'"
            self.curr.execute(s_sql2)
            self.res2 = self.curr.fetchall()

            if self.res2 is not None:
                
                return self.res2
                
        except Exception as e:
            logging.info('< Error > load_check_member_data_list2 : ' + str(e))

        finally:
            self.__disconnect__()

    ########################################
    # load_card_reader_member_list_detail
    ########################################
    def load_card_reader_member_list_detail(self , e_name):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            e_name = str(e_name).strip()

            s_sql = f"select e_name , p_name , r_time from card_reader_{now_day} where e_name='{e_name}' order by r_time asc"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
                
        except Exception as e:
            logging.info('< Error > load_card_reader_member_list_detail : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # load_card_reader_member_list2
    #################################
    def load_card_reader_member_list2(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep  = str(dep).strip()
            dep2 = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'

            s_sql = f"select distinct e_name from card_reader_{now_day} where d_name='{dep2}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchall()
            
            for val in res:
                
                check_name = str(val[0]).strip()
                s_sql2 = f"select distinct e_name from factory_hr_a where d_name ='{dep}'"
                self.curr.execute(s_sql2)
                res2 = self.curr.fetchone()
                    
                if val[0] == res2[0]:
                    return False
                else:
                    return res
                
        except Exception as e:
            logging.info('< Error > load_card_reader_member_list : ' + str(e))

        finally:
            self.__disconnect__()

    ###############################################
    # load_card_reader_member_check_status_list2
    ###############################################
    def load_card_reader_member_check_status_list2(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep  = str(dep).strip()
            dep2 = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select factory_hr_a.e_name from factory_hr_a inner join in_out_{now_day} on factory_hr_a.e_name!=in_out_{now_day}.e_name='{dep}' where factory_hr_a.d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchone()
            
            if res is not None:
                return res[0]  
                
        except Exception as e:
            logging.info('< Error > load_card_reader_member_list_real_total : ' + str(e))

        finally:
            self.__disconnect__()

    ############################################
    # load_card_reader_member_list_real_total
    ############################################
    def load_card_reader_member_list_real_total(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep  = str(dep).strip()
            dep2 = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select count(*) from in_out_{now_day} where d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchone()
            
            if res is not None:
                return res[0]  
                
        except Exception as e:
            logging.info('< Error > load_card_reader_member_list_real_total : ' + str(e))

        finally:
            self.__disconnect__()

    #######################################
    # load_card_reader_member_list_total
    #######################################
    def load_card_reader_member_list_total(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select count(*) from factory_hr_a where d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchone()
            
            if res is not None:
                return res[0]  
                
        except Exception as e:
            logging.info('< Error > load_card_reader_member_list_total : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##############################################
    # load_card_reader_member_check_status_list
    ##############################################
    def load_card_reader_member_check_status_list(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select distinct e_name from in_out_{now_day} where d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchall()
            
            return res
                
        except Exception as e:
            logging.info('< ERROR > load_card_reader_member_check_status_list : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # load_card_reader_member_list
    #################################
    def load_card_reader_member_list(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select distinct e_name from factory_hr_a where d_name='{dep}'"
            #s_sql = f"select distinct factory_hr_a.e_name from factory_hr_a inner join in_out_{now_day} on factory_hr_a.e_name = in_out_{now_day}.e_name where factory_hr_a.d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            for val in res:

                try:
                    a_sql =  f"create table in_out_{now_day}("
                    a_sql += f"no int not null primary key AUTO_INCREMENT,"
                    a_sql += f"r_date varchar(20) null,"
                    a_sql += f"r_time varchar(20) null,"
                    a_sql += f"d_id varchar(50) null,"
                    a_sql += f"d_name varchar(50) null,"
                    a_sql += f"e_id varchar(50) null,"
                    a_sql += f"e_name varchar(50) null,"
                    a_sql += f"p_id varchar(50) null,"
                    a_sql += f"p_name varchar(50) null,"
                    a_sql += f"c_id varchar(50) null"
                    a_sql += f")ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_unicode_ci;"

                    self.curr.execute(a_sql)

                except Exception as e:
                    
                    s_sql = f"select e_name from card_reader_{now_day} where e_name='{val[0]}' order by r_time desc limit 0,1"
                    self.curr.execute(s_sql)
                    res2 = self.curr.fetchone()

                    if res2 is not None:

                        s_sql3 = f"select e_name from in_out_{now_day} where d_name='{dep}' and e_name='{res2[0]}' order by no desc limit 0,1"
                        self.curr.execute(s_sql3)
                        res3 = self.curr.fetchone()

                        if res3 is None:

                            a_sql = f"insert into in_out_{now_day}(d_name , e_name) value('{dep}' , '{res2[0]}')"
                            self.curr.execute(a_sql)

                finally:
                    pass
            
            return res
                
        except Exception as e:
            logging.info('< ERROR > load_card_reader_member_list : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # load_group_member_list
    ###########################
    def load_group_member_list(self , e_name):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 
                
            s_sql = f"SELECT TRIM(d_name) FROM `card_reader_{now_day}` where e_name='{e_name}' order by d_name desc limit 0,1"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchone()

            s_sql2 = f"select distinct e_name from card_reader_{now_day} where d_name='{self.res[0]}'"
            self.curr.execute(s_sql2)
            self.res2 = self.curr.fetchall()

            return self.res2
                
        except Exception as e:
            logging.info('< Error > load_group_member_list : ' + str(e))

        finally:
            self.__disconnect__()

    ################################
    # load_check_member_data_list2
    ################################
    def load_check_member_data_list2(self , e_name):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 
                
            s_sql2 = f"SELECT d_name FROM `card_reader_20230915` group by d_name  order by d_name asc"
            self.curr.execute(s_sql2)
            self.res2 = self.curr.fetchall()

            if self.res2 is not None:
                return self.res2
                
        except Exception as e:
            logging.info('< Error > load_check_member_data_list2 : ' + str(e))

        finally:
            self.__disconnect__()

    ################################
    # load_check_member_data_list
    ################################
    def load_check_member_data_list(self , check_year , check_month , employee_name):
        
        self.__connect__()

        try:
            sql = f"select self_item_1_1 , self_item_1_s , self_item_1_3 , self_item_1_4 from check_member where check_year='{check_year}' and check_month='{check_month}' and employee_name='{employee_name}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res

        except Exception as e:
            logging.info('< Error > load_check_member_data_list : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # update_submit_check_member_2
    #################################
    def update_submit_check_member_2(self , employee_id , employee_name , check_year , check_month , sir_num1_1 , sir_num1_2 , sir_num1_3 , sir_num1_4 , sir_num2_1 , sir_num2_2 , sir_num2_3 , sir_num3_1 , sir_num3_2 , sir_num3_3 , sir_num4_1 , sir_num4_2 , sir_num4_3 , sir_num4_4 , sir_num5_1 , sir_num5_2 , sir_num5_3 , sir_num6_1 , sir_num6_2 , sir_num6_3 , sir_num7_1 , sir_num7_2 , sir_num7_3 , sir_num7_4 , sir_num8_1 , sir_num8_2 , sir_num8_3 , sir_num8_4 , sir_num8_5 , comment , other_total , sir_total , other_plus_total , final_total , final_comment):
        
        self.__connect__()

        try:
            sql = f"update check_member set sir_item_1_1='{sir_num1_1}' , sir_item_1_2='{sir_num1_2}' , sir_item_1_3='{sir_num1_3}' , sir_item_1_4='{sir_num1_4}' , sir_item_2_1='{sir_num2_1}' , sir_item_2_2='{sir_num2_2}' , sir_item_2_3='{sir_num2_3}' , sir_item_3_1='{sir_num3_1}' , sir_item_3_2='{sir_num3_2}' , sir_item_3_3='{sir_num3_3}' , sir_item_4_1='{sir_num4_1}' , sir_item_4_2='{sir_num4_2}' , sir_item_4_3='{sir_num4_3}' , sir_item_4_4='{sir_num4_4}' , sir_item_5_1='{sir_num5_1}' , sir_item_5_2='{sir_num5_2}' , sir_item_5_3='{sir_num5_3}' , sir_item_6_1='{sir_num6_1}' , sir_item_6_2='{sir_num6_2}' , sir_item_6_3='{sir_num6_3}' , sir_item_7_1='{sir_num7_1}' , sir_item_7_2='{sir_num7_2}' , sir_item_7_3='{sir_num7_3}' , sir_item_7_4='{sir_num7_4}' , sir_item_8_1='{sir_num8_1}' , sir_item_8_2='{sir_num8_2}' , sir_item_8_3='{sir_num8_3}' , sir_item_8_4='{sir_num8_4}' , sir_item_8_5='{sir_num8_5}' , comment='{comment}' , sir_total='{sir_total}' , other_total='{other_total}' , other_plus_total='{other_plus_total}' , final_total='{final_total}' , final_comment='{final_comment}' , sir_check='done' where employee_id='{employee_id}' and employee_name='{employee_name}' and check_year='{check_year}' and check_month='{check_month}'"
            self.res = self.curr.execute(sql)
            self.conn.commit()

            if self.res:
                b_res = 'ok'
                return b_res

        except Exception as e:
            logging.info('< Error > update_submit_check_member_2 : ' + str(e))

        finally:
            self.__disconnect__()
    
    #####################################
    # check_add_check_member_self_list
    #####################################
    def check_add_check_member_self_list(self):
        
        self.__connect__()

        try:
            
            sql = f"select check_year , check_month , employee_name from check_member where department_id='1BA' order by no"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('< Error > check_add_check_member_self_list : ' + str(e))

        finally:
            self.__disconnect__()

    #######################
    # search_member_item
    #######################
    def search_member_item(self , item , a_user):
        
        self.__connect__()

        try:
            
            '''
            conn_str        = f"DRIVER={{SQL Server}};SERVER={otsuka_factory2['host']};DATABASE={otsuka_factory2['db']};UID={otsuka_factory2['user']};PWD={otsuka_factory2['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql        = f"select {item} from T_HR_Employee where EmployeeName='{a_user}'"
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchone()

            return self.res[0]
            '''
            
            sql = f"select {item} from check_member where employee_name='{a_user}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is not None:
                return self.res[0]
        
        except Exception as e:
            logging.info('< Error > search_member_item : ' + str(e))

        finally:
            self.__disconnect__()
            #self.curr_mssql.close()
            #self.conn_mssql.close()

    #############################
    # show_work_time_total_val
    #############################
    def show_work_time_total_val(self , e_name , e_id , item):
        
        self.__connect__()

        try:
            
            s_sql = f"select sum({item}) from work_time where e_name='{e_name}' and e_id='{e_id}'"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchone()

            if self.res is not None:
                return self.res[0]
        
        except Exception as e:
            logging.info('< Error > show_work_time_total_val : ' + str(e))

        finally:
            self.__disconnect__()

    ######################## 
    # show_work_time_list
    ########################
    def show_work_time_list(self , e_name , e_id):
        
        self.__connect__()

        try:
            
            s_sql = f"select b_date from work_time where e_name='{e_name}' and e_id='{e_id}' order by b_date desc"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('< Error > show_work_time_list : ' + str(e))

        finally:
            self.__disconnect__()

    ##########################
    # submit_work_time_form
    ##########################
    def submit_work_time_form(self , a_work_no , a_name , dep_id , b_date , total_time , normal_time , over_time , availability_time , a_work_station_1 , a_production_1 , a_product_no_1 , a_work_normal_time_1 , a_work_over_time_1 , a_work_availability_time_1 , a_work_remark_1 , a_work_station_2 , a_production_2 , a_product_no_2 , a_work_normal_time_2 , a_work_over_time_2 , a_work_availability_time_2 , a_work_remark_2 , a_work_station_3 , a_production_3 , a_product_no_3 , a_work_normal_time_3 , a_work_over_time_3 , a_work_availability_time_3 , a_work_remark_3 , a_work_station_4 , a_production_4 , a_product_no_4 , a_work_normal_time_4 , a_work_over_time_4 , a_work_availability_time_4 , a_work_remark_4 , a_work_station_5 , a_production_5 , a_product_no_5 , a_work_normal_time_5 , a_work_over_time_5 , a_work_availability_time_5 , a_work_remark_5 , a_work_station_6 , a_production_6 , a_product_no_6 , a_work_normal_time_6 , a_work_over_time_6 , a_work_availability_time_6 , a_work_remark_6 , a_work_station_7 , a_production_7 , a_product_no_7 , a_work_normal_time_7 , a_work_over_time_7 , a_work_availability_time_7 , a_work_remark_7 , a_work_station_8 , a_production_8 , a_product_no_8 , a_work_normal_time_8 , a_work_over_time_8 , a_work_availability_time_8 , a_work_remark_8 , a_work_station_9 , a_production_9 , a_product_no_9 , a_work_normal_time_9 , a_work_over_time_9 , a_work_availability_time_9 , a_work_remark_9 , a_work_station_10 , a_production_10 , a_product_no_10 , a_work_normal_time_10 , a_work_over_time_10 , a_work_availability_time_10 , a_work_remark_10 , a_work_station_11 , a_production_11 , a_product_no_11 , a_work_normal_time_11 , a_work_over_time_11 , a_work_availability_time_11 , a_work_remark_11 , a_work_station_12 , a_production_12 , a_product_no_12 , a_work_normal_time_12 , a_work_over_time_12 , a_work_availability_time_12 , a_work_remark_12):
        
        self.__connect__()
        
        try:
        
            s_sql = f"select * from work_time where e_id='{a_work_no}' and e_name='{a_name}' and b_date='{b_date}' and dep_id='{dep_id}'"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchone()

            data  = b_date.split('-')
            r_year  = data[0]
            r_month = data[1]
            r_day   = data[2]
            
            if self.res is None:
                a_sql  = f"insert into work_time("
                a_sql += f"e_id , e_name , dep_id , b_date , total_time , normal_time , over_time , availability_time , r_year , r_month , r_day ," 
                a_sql += f"w_s_1 , w_s_1_product , w_s_1_num , w_s_1_normal_time , w_s_1_over_time , w_s_1_avail_time , w_s_1_remark , " 
                a_sql += f"w_s_2 , w_s_2_product , w_s_2_num , w_s_2_normal_time , w_s_2_over_time , w_s_2_avail_time , w_s_2_remark , " 
                a_sql += f"w_s_3 , w_s_3_product , w_s_3_num , w_s_3_normal_time , w_s_3_over_time , w_s_3_avail_time , w_s_3_remark , " 
                a_sql += f"w_s_4 , w_s_4_product , w_s_4_num , w_s_4_normal_time , w_s_4_over_time , w_s_4_avail_time , w_s_4_remark , " 
                a_sql += f"w_s_5 , w_s_5_product , w_s_5_num , w_s_5_normal_time , w_s_5_over_time , w_s_5_avail_time , w_s_5_remark , " 
                a_sql += f"w_s_6 , w_s_6_product , w_s_6_num , w_s_6_normal_time , w_s_6_over_time , w_s_6_avail_time , w_s_6_remark , " 
                a_sql += f"w_s_7 , w_s_7_product , w_s_7_num , w_s_7_normal_time , w_s_7_over_time , w_s_7_avail_time , w_s_7_remark , " 
                a_sql += f"w_s_8 , w_s_8_product , w_s_8_num , w_s_8_normal_time , w_s_8_over_time , w_s_8_avail_time , w_s_8_remark , " 
                a_sql += f"w_s_9 , w_s_9_product , w_s_9_num , w_s_9_normal_time , w_s_9_over_time , w_s_9_avail_time , w_s_9_remark , " 
                a_sql += f"w_s_10 , w_s_10_product , w_s_10_num , w_s_10_normal_time , w_s_10_over_time , w_s_10_avail_time , w_s_10_remark , " 
                a_sql += f"w_s_11 , w_s_11_product , w_s_11_num , w_s_11_normal_time , w_s_11_over_time , w_s_11_avail_time , w_s_11_remark , " 
                a_sql += f"w_s_12 , w_s_12_product , w_s_12_num , w_s_12_normal_time , w_s_12_over_time , w_s_12_avail_time , w_s_12_remark" 
                a_sql += f") value("
                a_sql += f"'{a_work_no}' , '{a_name}' , '{dep_id}' ,'{b_date}', '{total_time}' , '{normal_time}' , '{over_time}' , '{availability_time}' , '{r_year}' , '{r_month}' , '{r_day}' ,"
                a_sql += f"'{a_work_station_1}' , '{a_production_1}' , '{a_product_no_1}' , '{a_work_normal_time_1}' , '{a_work_over_time_1}' , '{a_work_availability_time_1}' , '{a_work_remark_1}' , "
                a_sql += f"'{a_work_station_2}' , '{a_production_2}' , '{a_product_no_2}' , '{a_work_normal_time_2}' , '{a_work_over_time_2}' , '{a_work_availability_time_2}' , '{a_work_remark_2}' , "
                a_sql += f"'{a_work_station_3}' , '{a_production_3}' , '{a_product_no_3}' , '{a_work_normal_time_3}' , '{a_work_over_time_3}' , '{a_work_availability_time_3}' , '{a_work_remark_3}' , "
                a_sql += f"'{a_work_station_4}' , '{a_production_4}' , '{a_product_no_4}' , '{a_work_normal_time_4}' , '{a_work_over_time_4}' , '{a_work_availability_time_4}' , '{a_work_remark_4}' , "
                a_sql += f"'{a_work_station_5}' , '{a_production_5}' , '{a_product_no_5}' , '{a_work_normal_time_5}' , '{a_work_over_time_5}' , '{a_work_availability_time_5}' , '{a_work_remark_5}' , "
                a_sql += f"'{a_work_station_6}' , '{a_production_6}' , '{a_product_no_6}' , '{a_work_normal_time_6}' , '{a_work_over_time_6}' , '{a_work_availability_time_6}' , '{a_work_remark_6}' , "
                a_sql += f"'{a_work_station_7}' , '{a_production_7}' , '{a_product_no_7}' , '{a_work_normal_time_7}' , '{a_work_over_time_7}' , '{a_work_availability_time_7}' , '{a_work_remark_7}' , "
                a_sql += f"'{a_work_station_8}' , '{a_production_8}' , '{a_product_no_8}' , '{a_work_normal_time_8}' , '{a_work_over_time_8}' , '{a_work_availability_time_8}' , '{a_work_remark_8}' , "
                a_sql += f"'{a_work_station_9}' , '{a_production_9}' , '{a_product_no_9}' , '{a_work_normal_time_9}' , '{a_work_over_time_9}' , '{a_work_availability_time_9}' , '{a_work_remark_9}' , "
                a_sql += f"'{a_work_station_10}' , '{a_production_10}' , '{a_product_no_10}' , '{a_work_normal_time_10}' , '{a_work_over_time_10}' , '{a_work_availability_time_10}' , '{a_work_remark_10}' , "
                a_sql += f"'{a_work_station_11}' , '{a_production_11}' , '{a_product_no_11}' , '{a_work_normal_time_11}' , '{a_work_over_time_11}' , '{a_work_availability_time_11}' , '{a_work_remark_11}' , "
                a_sql += f"'{a_work_station_12}' , '{a_production_12}' , '{a_product_no_12}' , '{a_work_normal_time_12}' , '{a_work_over_time_12}' , '{a_work_availability_time_12}' , '{a_work_remark_12}'"
                a_sql += f")" 

                res = self.curr.execute(a_sql)

            else:
                r_val = 'no'
                return r_val
        
        except Exception as e:
            logging.info('< Error > submit_work_time_form : ' + str(e))

        finally:
            self.__disconnect__()
    
    ################
    # search_item
    ################
    def search_item(self , item , a_user):
        
        self.__connect__()
        
        try:
            
            '''
            conn_str        = f"DRIVER={{SQL Server}};SERVER={otsuka_factory2['host']};DATABASE={otsuka_factory2['db']};UID={otsuka_factory2['user']};PWD={otsuka_factory2['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql        = f"select {item} from T_HR_Employee where EmployeeName='{a_user}'"
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchone()

            return self.res[0]
            '''
        
            sql = f"select {item} from hr_a where employee_name='{a_user}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()
            
            if self.res is not None:
                return self.res[0]
        
        except Exception as e:
            logging.info('< Error > search_item : ' + str(e))

        finally:
            self.__disconnect__()
            #self.curr_mssql.close()
            #self.conn_mssql.close()

    ################################
    # check_add_check_member_list
    ################################
    def check_add_check_member_list(self , employee_name):
        
        self.__connect__()
        
        try:
            
            sql = f"select check_year , check_month , employee_name from check_member where employee_name='{employee_name}' order by b_date"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
            
            else:
                self.res = '沒考核紀錄。'
                return self.res
            
        
        except Exception as e:
            logging.info('< Error > check_add_check_member_list : ' + str(e))

        finally:
            self.__disconnect__()
    
    #################################
    # check_add_check_member_data
    #################################
    def check_add_check_member_data(self , employee_name , check_year , check_month):
        try:
            self.__connect__()
            
            sql = f"select employee_name from check_member where employee_name='{employee_name}' and check_year='{check_year}' and check_month='{check_month}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is None:

                res_a = 'ok'
                return res_a
                
            else: 

                res_a = 'no'
                return res_a
        
        except Exception as e:
            logging.info('< Error > check_add_check_member_data : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # submit_add_check_member_data
    #################################
    def submit_add_check_member_data(self , employee_id , employee_name , department_id , department_name , job_title , b_date , end_date , check_year , check_month , self_num1_1 , self_num1_2 , self_num1_3 , self_num1_4 , self_num2_1 , self_num2_2 , self_num2_3 , self_num3_1 , self_num3_2 , self_num3_3 , self_num4_1 , self_num4_2 , self_num4_3 , self_num4_4 , self_num5_1 , self_num5_2 , self_num5_3 , self_num6_1 , self_num6_2 , self_num6_3  , self_total):
        
        self.__connect__()
        
        try:
            
            sql = f"select employee_name from check_member where employee_name='{employee_name}' and check_year='{check_year}' and check_month='{check_month}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is None:
                
                sql2 = f"insert into check_member(employee_id , employee_name , department_id , department_name , b_date , end_date , check_year , check_month , self_item_1_1 , self_item_1_2 , self_item_1_3 , self_item_1_4 , self_item_2_1 , self_item_2_2 , self_item_2_3 , self_item_3_1 , self_item_3_2 , self_item_3_3 , self_item_4_1 ,  self_item_4_2 ,  self_item_4_3 ,  self_item_4_4 ,  self_item_5_1 , self_item_5_2 , self_item_5_3 , self_item_6_1 , self_item_6_2 , self_item_6_3 , self_total , self_check) value('{employee_id}' , '{employee_name}' , '{department_id}' , '{department_name}' , '{b_date}' , '{end_date}' , '{check_year}' , '{check_month}' , '{self_num1_1}' , '{self_num1_2}' , '{self_num1_3}' , '{self_num1_4}' , '{self_num2_1}' , '{self_num2_2}' , '{self_num2_3}' , '{self_num3_1}' , '{self_num3_2}' , '{self_num3_3}' , '{self_num4_1}' , '{self_num4_2}' , '{self_num4_3}' , '{self_num4_4}' , '{self_num5_1}' , '{self_num5_2}' , '{self_num5_3}' , '{self_num6_1}' , '{self_num6_2}' , '{self_num6_3}'  , '{self_total}' , 'done')" 
                self.curr.execute(sql2)
                self.conn.commit()

        except Exception as e:
            logging.info('< Error > submit_add_check_member_data : ' + str(e))

        finally:
            self.__disconnect__() 
    
    #############################
    # submit_add_check_account
    #############################
    def submit_add_check_account(self , employee_id , employee_name , login_id , mobile , department_name , department_code , company_id , end_date):
        try:
            self.__connect__()
            
            sql = f"select employee_name from hr_a where employee_name='{employee_name}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is None:
                
                sql3 = f"select employee_name from hr_a where login_id='{login_id}'"
                self.curr.execute(sql3)
                self.res3 = self.curr.fetchone()

                if self.res3 is None:
                
                    sql2 = f"insert into hr_a(employee_id , employee_name , login_id , mobile , department_name , department_code , company_id , end_date) value('{employee_id}' , '{employee_name}' , '{login_id}' , '{mobile}' , '{department_name}' , '{department_code}' , '{company_id}' , '{end_date}')" 
                    self.curr.execute(sql2)

                    res_a = 'ok'
                    return res_a
                
                else:

                    res_a = 'no_login_id'
                    return res_a    
            
            else:

                res_a = 'no'
                return res_a
        
        except Exception as e:
            logging.info('< Error > submit_add_check_account : ' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # load_account_data_form_self_item
    #####################################
    def load_account_data_form_self_item(self , employee_id , employee_name , check_year , check_month):
        
        self.__connect__()

        try:
            
            sql = f"select self_item_1_1 , self_item_1_2 , self_item_1_3 , self_item_1_4 , self_item_2_1 , self_item_2_1 , self_item_2_3 , self_item_3_1 , self_item_3_2 , self_item_3_3 , self_item_4_1 , self_item_4_2 , self_item_4_3 , self_item_4_4 , self_item_5_1 , self_item_5_2 , self_item_5_3 , self_item_6_1 , self_item_6_2 , self_item_6_3 , self_total from check_member where employee_id='{employee_id}' and employee_name='{employee_name}' and check_year='{check_year}' and check_month='{check_month}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('< Error > load_account_data_form_self_item : ' + str(e))

        finally:
            self.__disconnect__()
    
    ################################
    # load_account_data_form_item
    ################################
    def load_account_data_form_item(self , item , employee_id):
        try:
            self.__connect__()
            
            sql = f"select {item}  from check_member where employee_id='{employee_id}' order by check_year desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('< Error > load_account_data_form_item : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # load_account_data_item
    ###########################
    def load_account_data_item(self , user):
        try:
            self.__connect__()
            
            sql = f"select employee_id , employee_name , end_date from hr_a where employee_id='{user}' and department_code like '1B%' and job_title_name != '經理' "
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('< Error > load_account_data_item : ' + str(e))

        finally:
            self.__disconnect__()

    ############################
    # factory_check_form_item
    ############################
    def factory_check_form_item(self , user):
        
        self.__connect__()
        
        try:
            
            sql = f"select job_title_name from hr_a where employee_name='{user}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()
            
            if self.res is not None:
                return self.res[0]
        
        except Exception as e:
            logging.info('< Error > factory_check_form_item : ' + str(e))

        finally:
            self.__disconnect__()

    ############################
    # factory_check_form_list
    ############################
    def factory_check_form_list(self):
        
        self.__connect__()
        
        try:
            
            sql = "select employee_name from hr_a where department_code like '1B%' and job_title_name != '經理' order by no desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('< Error > factory_check_form_list : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##############################
    # factory_work_account_list
    ##############################
    def factory_work_account_list(self):
        try:
            self.__connect__()
            
            sql = "select a_user , a_name , a_pwd , a_status , a_work_no from account where a_lv='3' and a_position='生二部' order by no desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('< Error > factory_work_account_list : ' + str(e))

        finally:
            self.__disconnect__()

    ################
    # add_account
    ################
    def add_account(self , a_date , a_name , a_work_no , a_position , a_status , a_user):
        try:
            self.__connect__()
            
            ### time record
            now_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            r_year   = time.strftime("%Y" , time.localtime()) 
            r_month  = time.strftime("%m" , time.localtime()) 
            r_day    = time.strftime("%d" , time.localtime()) 
            r_time   = time.strftime("%H:%M:%S" , time.localtime()) 
            a_pwd    = 'm' + a_work_no
            a_lv     = 3

            sql = "select a_user from account where a_user='{0}'".format(a_user)
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is None:
                
                if a_status == '使用':
                    a_status = 'run'
                else:
                    a_status = 'stop'

                add_sql  = "insert into account(r_year , r_month , r_day , r_time , a_name , a_pwd , a_work_no , a_position , a_status , a_lv , a_user)"
                add_sql += " value ('{0}' , '{1}' ,'{2}' ,'{3}' ,'{4}' ,'{5}' ,'{6}' ,'{7}' ,'{8}' ,'{9}' , '{10}')".format(r_year , r_month , r_day , r_time , a_name , a_pwd , a_work_no , a_position , a_status , a_lv , a_user)

                self.curr.execute(add_sql)
                return True
                
            else:
                return False
        
        except Exception as e:
            logging.info('< Error > add_account : ' + str(e))

        finally:
            self.__disconnect__()

    #########################
    # factory_work_position
    #########################
    def factory_work_position(self):
        try:
            self.__connect__()
            
            sql = "select distinct c_content from work_position order by e_name desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('< Error > factory_work_position : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # factory_work_station_3
    ###########################
    def factory_work_station_3(self):
        
        self.__connect__()
        
        try:
            
            sql = "select distinct c_content from work_station_3 order by e_name desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('< Error > factory_work_station_3 : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # factory_work_station_1
    ###########################
    def factory_work_station_1(self):
        try:
            self.__connect__()
            
            sql = "select distinct c_content from work_station_1 order by e_name desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('< Error > factory_work_statio_1 : ' + str(e))

        finally:
            self.__disconnect__()

    #########################
    # factory_work_station
    #########################
    def factory_work_station(self):
        try:
            self.__connect__()
            
            sql = "select distinct c_content from work_station order by e_name desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('< Error > factory_work_station : ' + str(e))

        finally:
            self.__disconnect__()

    #####################
    # check_login_code
    #####################
    def check_login_code(self,user,login_code):
        
        try:
            self.user = user
            self.login_code = login_code

            self.__connect__()

            sql = "select login_code from login_out_record where a_user='{0}' order by no desc limit 0,1".format(self.user)
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res[0] == self.login_code:
                return 'ok'

        except Exception as e:
            logging.info("< Error > check login code : " + str(e))

        finally:
            self.__disconnect__()

    ###########
    # dep_id
    ###########
    def dep_id(self,user,pwd):
        
        try:
            self.login_id = user
            self.mobile   = pwd
            '''
            conn_str = f"DRIVER={{SQL Server}};SERVER={otsuka_factory2['host']};DATABASE={otsuka_factory2['db']};UID={otsuka_factory2['user']};PWD={otsuka_factory2['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql = f"select DepartmentID from T_HR_Employee where loginID='{self.login_id}' and Mobile='{self.mobile}'"
            self.curr_mssql.execute(self.sql)
            self.res = self.curr_mssql.fetchone()

            self.sql2 = f"select UpperDepartmentID from T_HR_Department where DepartmentID='{self.res[0]}'"
            self.curr_mssql.execute(self.sql2)
            self.res2 = self.curr_mssql.fetchone()

            return self.res2
            '''
            
            self.__connect__()

            self.sql = "select department_code from hr_a where login_id='{0}' and mobile='{1}'".format(self.login_id , self.mobile)
            self.curr.execute(self.sql)
            self.res        = self.curr.fetchone()
            self.dep_code   = self.res[0]
            self.r_dep_code = self.dep_code[0:2]

            ###########
            # 生一部
            ###########
            if self.r_dep_code == '1A':
                return self.r_dep_code
            ###########
            # 生二部
            ###########
            elif self.r_dep_code == '1B':
                return self.r_dep_code
            ###########
            # 生三部
            ###########
            elif self.r_dep_code == '1K':
                return self.r_dep_code
            ##############################
            # 剩下都依照原本部門代號顯示
            ##############################
            else:
                return self.dep_code

        except Exception as e:
            logging.error("< Error > dep_id : " + str(e))

        finally:
            self.__disconnect__()

    ##########
    # login
    ##########
    def login(self,user,pwd):
        
        try:
            self.login_id = user
            self.mobile   = pwd

            #########################
            #
            # connect MsSQL - SHRM
            #
            #########################
            '''
            conn_str = f"DRIVER={{SQL Server}};SERVER={otsuka_factory2['host']};DATABASE={otsuka_factory2['db']};UID={otsuka_factory2['user']};PWD={otsuka_factory2['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql = f"select EmployeeName from T_HR_Employee where loginID='{self.login_id}' and Mobile='{self.mobile}'"
            self.curr_mssql.execute(self.sql)
            self.res = self.curr_mssql.fetchone()

            return self.res
            '''
            
            #########################
            #
            # connect MySQL - hr_a
            #
            #########################
            self.__connect__()

            self.sql = f"select Employee_name from hr_a where login_id='{self.login_id}' and mobile='{self.mobile}'"
            self.curr.execute(self.sql)
            self.res = self.curr.fetchone()

            return self.res

        except Exception as e:
            logging.info("< Error > login : " + str(e))

        finally:
            self.__disconnect__()
        
    #################
    # login_record   
    ################# 
    def login_record(self,user,login_code,r_time,ip):
        
        try:
            self.user       = user
            self.login_code = login_code
            self.r_time     = r_time
            self.ip         = ip

            self.__connect__()

            self.sql2 = "insert into login_out_record(a_user,login_code,login_time,login_ip) value('{0}','{1}','{2}','{3}')".format(self.user , self.login_code , self.r_time , self.ip)
            self.curr.execute(self.sql2)

        except Exception as e:
            logging.info("< Error > login record : " + str(e))

        finally:
            self.__disconnect__()
    
    #####################
    # operation_record
    #####################
    def operation_record(self,r_time,user,login_code,item):
        
        self.__connect__()
        
        try:
            self.r_time     = r_time
            self.user       = user
            self.item       = item
            self.login_code = login_code
            
            self.sql = "insert into operation_record(r_time,a_user,item,login_code) value('{0}','{1}','{2}','{3}')".format(self.r_time , self.user , self.item , self.login_code)
            self.curr.execute(self.sql)

        except Exception as e:
            logging.info("< Error > operation record : " + str(e))

        finally:
            self.__disconnect__()
    
    ##################
    # logout_record
    ##################
    def logout_record(self,user,login_code,r_time):
        
        try:
            self.user = user
            self.login_code = login_code
            self.r_time = r_time

            self.__connect__()    

            self.sql = "update login_out_record set logout_time='{0}' where login_code='{1}' and a_user='{2}'".format(self.r_time , self.login_code , self.user)
            self.curr.execute(self.sql)

        except Exception as e:
            logging.info("< Error > logout record : " + str(e))

        finally:
            self.__disconnect__()

    ######################
    # __connect_mssql__ 
    ######################
    def __connect_mssql__(self):
        
        try:
            conn_str = f"DRIVER={{SQL Server}};SERVER={otsuka_factory2['host']};DATABASE={otsuka_factory2['db']};UID={otsuka_factory2['user']};PWD={otsuka_factory2['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            
        except Exception as e:
            logging.info("< ERROR > __connect_mssql__ " + str(e))

        finally:
            self.curr.close()
            self.conn_mssql.close()
    
    #########################
    # __disconnect_mssql__ 
    #########################
    def __disconnect_mssql__(self):
        
        try:
            self.curr_mssql.close()
            self.conn_mssql.close()
            
        except Exception as e:
            logging.info("< ERROR > __disconnect_mssql__ " + str(e))

        finally:
            pass

    ################
    # __connect__ 
    ################
    def __connect__(self):
        
        try:
            self.conn = pymysql.connect(host=otsuka_factory['host'],port=otsuka_factory['port'],user=otsuka_factory['user'],password=otsuka_factory['pwd'],database=otsuka_factory['db'],charset=otsuka_factory['charset'])
            self.curr = self.conn.cursor()

        except Exception as e:
            logging.info("< ERROR > __connect__ " + str(e))

        finally:
            pass

    ###################
    # __disconnect__
    ###################
    def __disconnect__(self):
        
        try:
            self.conn.commit()
            self.conn.close()

        except Exception as e:
            logging.info("< ERROR > __disconnect__ : " + str(e))

        finally:
            pass

    ################
    # __connect4__ 
    ################
    def __connect4__(self):
        
        try:
            self.conn = pymysql.connect(host=otsuka_factory4['host'],port=otsuka_factory4['port'],user=otsuka_factory4['user'],password=otsuka_factory4['pwd'],database=otsuka_factory4['db'],charset=otsuka_factory4['charset'])
            self.curr = self.conn.cursor()

        except Exception as e:
            logging.info("< ERROR > __connect4__ " + str(e))

        finally:
            pass

    ###################
    # __disconnect4__
    ###################
    def __disconnect4__(self):
        
        try:
            self.conn.commit()
            self.conn.close()

        except Exception as e:
            logging.info("< ERROR > __disconnect4__ : " + str(e))

        finally:
            pass

